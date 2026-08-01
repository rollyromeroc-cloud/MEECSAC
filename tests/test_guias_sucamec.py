from io import BytesIO
from zipfile import ZipFile

import openpyxl
from PIL import Image as ImagenPIL

from core.constants import PLANTILLA_GUIA_TIPO1, PLANTILLA_GUIA_TIPO2
from core.guias_sucamec import (
    RAIZ_PLANTILLAS,
    ProductoGuia,
    desglosar_guias,
    generar_zip_guias,
)
from core.models import PolvorinGuiaSucamec

POLVORIN_PRUEBA = PolvorinGuiaSucamec(
    nombre="Polvorín N.° 1",
    concesion_nombre="CONCESION DE PRUEBA",
    concesion_codigo="99999999X01",
    concesion_distrito="DISTRITO CONCESION",
    concesion_provincia="PROVINCIA CONCESION",
    concesion_departamento="DEPARTAMENTO CONCESION",
    resolucion_gerencia_numero="01463-2026-SUCAMEC/DEPP-SDAEPP",
    resolucion_gerencia_fecha="12/05/2026",
)


def _valor_original(ruta_plantilla, celda):
    wb = openpyxl.load_workbook(RAIZ_PLANTILLAS / ruta_plantilla)
    return wb.active[celda].value


def test_desglosar_guias_emulsion_con_restante():
    # Mismo caso dorado que test_polvorin.py: 8 completas de 575 kg + 1 restante de 225 kg.
    guias = desglosar_guias(cantidad_solicitada=4825, capacidad_por_guia=575)
    assert len(guias) == 9
    assert all(g.cantidad == 575 and g.tipo == "completa" for g in guias[:8])
    assert guias[8].cantidad == 225
    assert guias[8].tipo == "restante"


def test_desglosar_guias_exacta_sin_restante():
    guias = desglosar_guias(cantidad_solicitada=1150, capacidad_por_guia=575)
    assert len(guias) == 2
    assert all(g.tipo == "completa" for g in guias)


def _producto_emulsion(nombre_variante: str) -> ProductoGuia:
    return ProductoGuia(
        nombre_variante=nombre_variante,
        categoria="Explosivos",
        producto_sucamec="Emulsión o hidrogel encartuchada",
        cantidad_solicitada=4825,
        capacidad_por_guia=575,
        unidad_abrev="KG",
        polvorin=POLVORIN_PRUEBA,
    )


def test_generar_zip_guias_dos_variantes_no_se_combinan():
    # Ej. del usuario: Emulnor 3000 y Emulnor 5000, cada uno 9 guias (18 en total).
    productos = [_producto_emulsion("Emulnor 3000"), _producto_emulsion("Emulnor 5000")]
    zip_buffer, resumen = generar_zip_guias(productos)
    assert len(resumen) == 18

    direccion_original_t1 = _valor_original(PLANTILLA_GUIA_TIPO1["Explosivos"], "E51")
    resolucion_original_t1 = _valor_original(PLANTILLA_GUIA_TIPO1["Explosivos"], "S59")
    direccion_original_t2 = _valor_original(PLANTILLA_GUIA_TIPO2, "E40")

    with ZipFile(BytesIO(zip_buffer.getvalue())) as zf:
        nombres = zf.namelist()
        assert len(nombres) == 36  # 18 guias x (tipo 1 + tipo 2)

        nombre_g1 = next(n for n in nombres if n.startswith("TIPO1_Emulnor_3000_guia_01_completa"))
        wb = openpyxl.load_workbook(BytesIO(zf.read(nombre_g1)))
        ws = wb.active
        assert ws["I21"].value == "EMULSIÓN O HIDROGEL ENCARTUCHADA"
        assert ws["AG21"].value == 575
        assert ws["AL21"].value == "KILOGRAMOS"
        # La direccion/resolucion del polvorin (destino en tipo 1) NO se toca:
        # debe seguir siendo exactamente la del Excel base.
        assert ws["E51"].value == direccion_original_t1
        assert ws["S59"].value == resolucion_original_t1
        assert len(ws._images) == 2  # logos del modelo preservados

        nombre_restante = next(n for n in nombres if n.startswith("TIPO1_Emulnor_3000_guia_09_restante"))
        wb_restante = openpyxl.load_workbook(BytesIO(zf.read(nombre_restante)))
        assert wb_restante.active["AG21"].value == 225

        nombre_tipo2 = next(n for n in nombres if n.startswith("TIPO2_Emulnor_5000_guia_01_completa"))
        wb2 = openpyxl.load_workbook(BytesIO(zf.read(nombre_tipo2)))
        ws2 = wb2.active
        # El origen (polvorin) en tipo 2 tampoco se toca.
        assert ws2["E40"].value == direccion_original_t2
        # El destino (concesion) en tipo 2 si se completa con lo asignado.
        assert "CONCESION DE PRUEBA" in ws2["E51"].value
        assert "99999999X01" in ws2["E51"].value
        assert ws2["E55"].value == POLVORIN_PRUEBA.concesion_distrito
        # La resolucion de gerencia de la solicitud (bloque destino) tambien
        # se completa; la de subdireccion del polvorin (bloque origen) no.
        assert ws2["S59"].value == "N° 01463-2026-SUCAMEC/DEPP-SDAEPP"
        assert (ws2["AE59"].value, ws2["AH59"].value, ws2["AK59"].value) == (12, 5, 2026)
        resolucion_origen_original = _valor_original(PLANTILLA_GUIA_TIPO2, "S48")
        assert ws2["S48"].value == resolucion_origen_original


def test_resolucion_gerencia_vacia_no_sobreescribe_bloque_destino():
    # Si la solicitud no trae resolucion de gerencia, el bloque destino
    # (limpiado en la plantilla) se queda vacio en vez de heredar algo.
    polvorin_sin_resolucion = PolvorinGuiaSucamec(
        nombre="Solicitud sin resolución",
        concesion_nombre="OTRA CONCESION",
        concesion_codigo="11111111X01",
    )
    producto = ProductoGuia(
        nombre_variante="Dinamita sin resolución",
        categoria="Explosivos",
        producto_sucamec="Dinamita",
        cantidad_solicitada=575,
        capacidad_por_guia=575,
        unidad_abrev="KG",
        polvorin=polvorin_sin_resolucion,
    )
    zip_buffer, _ = generar_zip_guias([producto])
    with ZipFile(BytesIO(zip_buffer.getvalue())) as zf:
        nombre_tipo2 = next(n for n in zf.namelist() if n.startswith("TIPO2_"))
        wb = openpyxl.load_workbook(BytesIO(zf.read(nombre_tipo2)))
        assert wb.active["S59"].value is None


def test_generar_zip_guias_categoria_accesorios_usa_plantilla_correcta():
    producto = ProductoGuia(
        nombre_variante="Detonador X",
        categoria="Accesorios",
        producto_sucamec="Detonador de mecha o fulminante común",
        cantidad_solicitada=2500,
        capacidad_por_guia=500000,
        unidad_abrev="PZAS",
        polvorin=POLVORIN_PRUEBA,
    )
    zip_buffer, resumen = generar_zip_guias([producto])
    assert len(resumen) == 1  # menor a una capacidad -> 1 sola guia restante

    direccion_original_accesorios = _valor_original(PLANTILLA_GUIA_TIPO1["Accesorios"], "E51")

    with ZipFile(BytesIO(zip_buffer.getvalue())) as zf:
        nombre_tipo1 = next(n for n in zf.namelist() if n.startswith("TIPO1_Detonador_X"))
        wb = openpyxl.load_workbook(BytesIO(zf.read(nombre_tipo1)))
        ws = wb.active
        assert ws["I21"].value == "DETONADOR DE MECHA O FULMINANTE COMÚN"
        assert ws["AG21"].value == 2500
        assert ws["AL21"].value == "UNIDADES"
        # Categoria Accesorios -> debe usar la plantilla de accesorios (con su propia
        # direccion de destino), no la de explosivos.
        assert ws["E51"].value == direccion_original_accesorios


def _bytes_plantilla_modificada(ruta_relativa, celda, valor_nuevo) -> bytes:
    wb = openpyxl.load_workbook(RAIZ_PLANTILLAS / ruta_relativa)
    wb.active[celda] = valor_nuevo
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_polvorin_con_plantilla_propia_sobre_escribe_la_default():
    plantilla_propia = _bytes_plantilla_modificada(
        PLANTILLA_GUIA_TIPO1["Explosivos"], "E51", "DIRECCION DEL SEGUNDO POLVORIN",
    )
    polvorin_2 = PolvorinGuiaSucamec(
        nombre="Polvorín N.° 2",
        plantilla_tipo1_explosivos=plantilla_propia,
    )
    producto = ProductoGuia(
        nombre_variante="Dinamita polvorin 2",
        categoria="Explosivos",
        producto_sucamec="Dinamita",
        cantidad_solicitada=575,
        capacidad_por_guia=575,
        unidad_abrev="KG",
        polvorin=polvorin_2,
    )
    zip_buffer, _ = generar_zip_guias([producto])
    with ZipFile(BytesIO(zip_buffer.getvalue())) as zf:
        nombre_tipo1 = next(n for n in zf.namelist() if n.startswith("TIPO1_"))
        wb = openpyxl.load_workbook(BytesIO(zf.read(nombre_tipo1)))
        assert wb.active["E51"].value == "DIRECCION DEL SEGUNDO POLVORIN"


def _posicion_firma_original(ruta_relativa) -> tuple[int, int, int, int]:
    wb = openpyxl.load_workbook(RAIZ_PLANTILLAS / ruta_relativa)
    ws = wb.active
    firma = next(img for img in ws._images if img.anchor._from.row >= 50)
    return firma.anchor._from.col, firma.anchor._from.row, firma.width, firma.height


def test_sin_firma_nueva_se_conserva_la_del_modelo():
    # Los 3 modelos ya traen la firma del representante legal incrustada
    # (ademas del logo de SUCAMEC arriba); sin firma nueva, no se debe tocar.
    producto = _producto_emulsion("Emulnor 3000")
    zip_buffer, _ = generar_zip_guias([producto])  # sin firma_bytes
    with ZipFile(BytesIO(zip_buffer.getvalue())) as zf:
        nombre = next(n for n in zf.namelist() if n.startswith("TIPO1_Emulnor_3000_guia_01_completa"))
        wb = openpyxl.load_workbook(BytesIO(zf.read(nombre)))
        assert len(wb.active._images) == 2  # logo + firma original, intactos


def test_generar_zip_guias_reemplaza_la_firma_en_la_misma_posicion():
    firma = ImagenPIL.new("RGB", (400, 120), color=(10, 10, 10))
    buffer_firma = BytesIO()
    firma.save(buffer_firma, format="PNG")

    producto = _producto_emulsion("Emulnor 3000")
    zip_buffer, _ = generar_zip_guias([producto], firma_bytes=buffer_firma.getvalue())

    posiciones_originales = {
        "TIPO1_": _posicion_firma_original(PLANTILLA_GUIA_TIPO1["Explosivos"]),
        "TIPO2_": _posicion_firma_original(PLANTILLA_GUIA_TIPO2),
    }

    with ZipFile(BytesIO(zip_buffer.getvalue())) as zf:
        for tipo, (columna_original, fila_original, ancho_original, alto_original) in posiciones_originales.items():
            nombre = next(n for n in zf.namelist() if n.startswith(f"{tipo}Emulnor_3000_guia_01_completa"))
            wb = openpyxl.load_workbook(BytesIO(zf.read(nombre)))
            ws = wb.active
            # Sigue habiendo exactamente 2 imagenes: el logo (intacto) y la
            # firma nueva reemplazando a la original, en la misma posicion.
            assert len(ws._images) == 2
            imagenes_abajo = [img for img in ws._images if img.anchor._from.row >= 50]
            assert len(imagenes_abajo) == 1, f"Se esperaba 1 sola firma en {nombre}"
            imagen_firma = imagenes_abajo[0]
            assert imagen_firma.anchor._from.col == columna_original
            assert imagen_firma.anchor._from.row == fila_original
            # La firma no debe deformarse ni desbordar la caja original (misma escala).
            assert imagen_firma.width <= ancho_original
            assert imagen_firma.height <= alto_original
