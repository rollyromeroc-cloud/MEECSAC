"""Generador de guias SUCAMEC (Formato N.° 23) tipo 1 (FAMESA -> Polvorin) y
tipo 2 (Polvorin -> Unidad minera), a partir del calculo de guias por
producto (core.polvorin.calcular_guias) y de los datos de la concesion
asignados.

Las plantillas (bundled en assets/plantillas_sucamec/, o subidas por el
usuario por polvorin) son los modelos reales ya llenados: se abren tal cual y
solo se sobreescriben las celdas de producto/cantidad/unidades, la firma (si
se sube una imagen) y, en tipo 2, la concesion/unidad minera de destino. La
resolucion, direccion y demas datos propios del polvorin (destino en tipo 1,
origen en tipo 2) NO se tocan: se mantienen exactamente como estan en el
Excel base de ese polvorin.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl.drawing.image import Image as ImagenExcel
from openpyxl.utils import get_column_letter
from PIL import Image as ImagenPIL

from core.constants import PLANTILLA_GUIA_TIPO1, PLANTILLA_GUIA_TIPO2, UNIDAD_SUCAMEC_TEXTO
from core.models import PolvorinGuiaSucamec
from core.polvorin import calcular_guias

RAIZ_PLANTILLAS = Path(__file__).resolve().parent.parent / "assets" / "plantillas_sucamec"

# Los 3 modelos que envio el usuario ya traen la firma del representante
# legal como imagen incrustada (ej. "Formato gts TIPO 2.xlsx"), ademas del
# logo de SUCAMEC arriba de todo (filas 1-4). La firma se reconoce por estar
# mas abajo en la hoja; si no se sube una firma nueva, se deja la del modelo
# tal cual. Si se sube una, se quita la del modelo y se pone la nueva
# exactamente en su misma posicion y tamaño.
FILA_MINIMA_FIRMA = 50
FIRMA_ANCHO_MAX_PX_DEFECTO = 200
FIRMA_ALTO_MAX_PX_DEFECTO = 70


@dataclass
class GuiaIndividual:
    numero: int
    cantidad: float
    tipo: str  # "completa" o "restante"


@dataclass
class ProductoGuia:
    """Un producto/variante con su cantidad solicitada y el polvorin asignado."""

    nombre_variante: str
    categoria: str  # "Explosivos" o "Accesorios"
    producto_sucamec: str
    cantidad_solicitada: float
    capacidad_por_guia: float
    unidad_abrev: str  # "KG", "M" o "PZAS"
    polvorin: PolvorinGuiaSucamec


def desglosar_guias(cantidad_solicitada: float, capacidad_por_guia: float) -> list[GuiaIndividual]:
    """Convierte el resultado de calcular_guias en la lista de guias
    individuales a generar (N completas con la capacidad maxima por guia, mas
    1 restante si sobra una cantidad menor a la capacidad)."""
    resultado = calcular_guias(cantidad_solicitada, capacidad_por_guia)
    guias: list[GuiaIndividual] = []
    numero = 1
    for _ in range(resultado["guias_completas"]):
        guias.append(GuiaIndividual(numero=numero, cantidad=capacidad_por_guia, tipo="completa"))
        numero += 1
    if resultado["guia_restante"]:
        guias.append(GuiaIndividual(numero=numero, cantidad=resultado["cantidad_restante"], tipo="restante"))
    return guias


def _rellenar_producto(ws, producto: str, cantidad: float, unidad_abrev: str) -> None:
    ws["I21"] = producto.upper()
    ws["AG21"] = cantidad
    ws["AL21"] = UNIDAD_SUCAMEC_TEXTO.get(unidad_abrev, unidad_abrev)


def _rellenar_destino_concesion_tipo2(ws, polvorin: PolvorinGuiaSucamec) -> None:
    if polvorin.concesion_nombre or polvorin.concesion_codigo:
        ws["E51"] = (
            f"CONCESIÓN MINERA {polvorin.concesion_nombre} CON CODIGO UNICO:\n"
            f"{polvorin.concesion_codigo}\n"
        )
    if polvorin.concesion_distrito:
        ws["E55"] = polvorin.concesion_distrito
    if polvorin.concesion_provincia:
        ws["S55"] = polvorin.concesion_provincia
    if polvorin.concesion_departamento:
        ws["AE55"] = polvorin.concesion_departamento


def _preparar_imagen_firma(firma_bytes: bytes, ancho_max_px: int, alto_max_px: int) -> BytesIO:
    """Reescala la firma para que quepa en el recuadro (mismo tamaño que la
    firma original del modelo) sin deformarla."""
    imagen = ImagenPIL.open(BytesIO(firma_bytes))
    if imagen.mode not in ("RGB", "RGBA"):
        imagen = imagen.convert("RGBA")
    escala = min(ancho_max_px / imagen.width, alto_max_px / imagen.height)
    nuevo_ancho = max(1, round(imagen.width * escala))
    nuevo_alto = max(1, round(imagen.height * escala))
    imagen = imagen.resize((nuevo_ancho, nuevo_alto))
    salida = BytesIO()
    imagen.save(salida, format="PNG")
    salida.seek(0)
    return salida


def _quitar_firma_del_modelo(ws) -> tuple[int, int, int, int]:
    """Quita del modelo la imagen de firma ya incrustada (la que no es el logo
    de SUCAMEC, reconocida por estar mas abajo en la hoja) y devuelve su
    posicion y tamaño (columna, fila, ancho_px, alto_px) para reutilizarlos."""
    for imagen in list(ws._images):
        fila = imagen.anchor._from.row
        if fila >= FILA_MINIMA_FIRMA:
            columna = imagen.anchor._from.col
            ancho, alto = imagen.width, imagen.height
            ws._images.remove(imagen)
            return columna, fila, ancho, alto
    # Ningun modelo debería llegar aquí (los 3 traen firma), pero por si acaso
    # se define una posición y tamaño de respaldo razonables.
    return 15, 91, FIRMA_ANCHO_MAX_PX_DEFECTO, FIRMA_ALTO_MAX_PX_DEFECTO


def _insertar_firma(ws, firma_bytes: bytes | None) -> None:
    if not firma_bytes:
        return  # sin firma nueva, se deja la del modelo tal cual.
    columna, fila, ancho_max, alto_max = _quitar_firma_del_modelo(ws)
    buffer_imagen = _preparar_imagen_firma(firma_bytes, ancho_max, alto_max)
    imagen_pil = ImagenPIL.open(buffer_imagen)
    imagen_excel = ImagenExcel(buffer_imagen)
    imagen_excel.width, imagen_excel.height = imagen_pil.size
    celda = f"{get_column_letter(columna + 1)}{fila + 1}"
    ws.add_image(imagen_excel, celda)


def _cargar_plantilla(datos_plantilla: bytes | None, ruta_por_defecto: Path):
    origen = BytesIO(datos_plantilla) if datos_plantilla else ruta_por_defecto
    return openpyxl.load_workbook(origen)


def generar_guia_tipo1(producto: ProductoGuia, guia: GuiaIndividual, firma_bytes: bytes | None = None) -> BytesIO:
    plantilla_polvorin = (
        producto.polvorin.plantilla_tipo1_explosivos
        if producto.categoria == "Explosivos"
        else producto.polvorin.plantilla_tipo1_accesorios
    )
    ruta_por_defecto = RAIZ_PLANTILLAS / PLANTILLA_GUIA_TIPO1[producto.categoria]
    wb = _cargar_plantilla(plantilla_polvorin, ruta_por_defecto)
    ws = wb.active
    _rellenar_producto(ws, producto.producto_sucamec, guia.cantidad, producto.unidad_abrev)
    _insertar_firma(ws, firma_bytes)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_guia_tipo2(producto: ProductoGuia, guia: GuiaIndividual, firma_bytes: bytes | None = None) -> BytesIO:
    ruta_por_defecto = RAIZ_PLANTILLAS / PLANTILLA_GUIA_TIPO2
    wb = _cargar_plantilla(producto.polvorin.plantilla_tipo2, ruta_por_defecto)
    ws = wb.active
    _rellenar_producto(ws, producto.producto_sucamec, guia.cantidad, producto.unidad_abrev)
    _rellenar_destino_concesion_tipo2(ws, producto.polvorin)
    _insertar_firma(ws, firma_bytes)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _nombre_base(nombre_variante: str) -> str:
    limpio = "".join(caracter if caracter.isalnum() else "_" for caracter in nombre_variante.strip())
    limpio = "_".join(filter(None, limpio.split("_")))
    return limpio or "producto"


def generar_zip_guias(productos: list[ProductoGuia], firma_bytes: bytes | None = None) -> tuple[BytesIO, list[dict]]:
    """Genera un .zip con los Excel tipo 1 y tipo 2 de cada guia individual de
    cada producto/variante. Devuelve el .zip en memoria y un resumen de filas
    (para mostrar una tabla en la UI) con producto, guia, cantidad y archivos."""
    buffer_zip = BytesIO()
    resumen: list[dict] = []
    with ZipFile(buffer_zip, "w", ZIP_DEFLATED) as zf:
        for producto in productos:
            guias = desglosar_guias(producto.cantidad_solicitada, producto.capacidad_por_guia)
            base = _nombre_base(producto.nombre_variante)
            for guia in guias:
                sufijo = f"guia_{guia.numero:02d}_{guia.tipo}"

                buffer1 = generar_guia_tipo1(producto, guia, firma_bytes)
                nombre1 = f"TIPO1_{base}_{sufijo}.xlsx"
                zf.writestr(nombre1, buffer1.getvalue())

                buffer2 = generar_guia_tipo2(producto, guia, firma_bytes)
                nombre2 = f"TIPO2_{base}_{sufijo}.xlsx"
                zf.writestr(nombre2, buffer2.getvalue())

                resumen.append({
                    "Producto/variante": producto.nombre_variante,
                    "Polvorín": producto.polvorin.nombre,
                    "N.° de guía": guia.numero,
                    "Tipo de guía": guia.tipo,
                    "Cantidad": guia.cantidad,
                    "Unidad": producto.unidad_abrev,
                    "Archivo TIPO 1": nombre1,
                    "Archivo TIPO 2": nombre2,
                })
    buffer_zip.seek(0)
    return buffer_zip, resumen
